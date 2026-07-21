from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required
from datetime import date, timedelta
from decimal import Decimal

from app.extensions import db
from app.models.models import Party, PartyType, WorkOrder, Treatment, TreatmentCategory, ExchangeRate, Makbuz
from app.authz import permissions_required

parties_bp = Blueprint("parties", __name__)

MONTHS = [
    (1, "Ocak"), (2, "Şubat"), (3, "Mart"), (4, "Nisan"),
    (5, "Mayıs"), (6, "Haziran"), (7, "Temmuz"), (8, "Ağustos"),
    (9, "Eylül"), (10, "Ekim"), (11, "Kasım"), (12, "Aralık"),
]


def _get_treatments_by_category(category):
    from app.services.search_service import tr_order

    treatments = db.session.execute(
        db.select(Treatment)
        .where(Treatment.category == category, Treatment.is_active == True)
        .order_by(tr_order(Treatment.name))
    ).scalars().all()
    return [{"id": t.id, "name": t.name, "price": float(t.price_eur), "currency": t.currency} for t in treatments]


def _get_current_rate():
    rate = db.session.execute(
        db.select(ExchangeRate).order_by(ExchangeRate.rate_date.desc()).limit(1)
    ).scalar_one_or_none()
    return float(rate.eur_to_try) if rate else 0


@parties_bp.route("/")
@login_required
@permissions_required("clinical.view")
def list_parties():
    search = request.args.get("search", "").strip()

    query = db.select(Party).where(
        Party.party_type == PartyType.DENTIST,
        Party.is_active == True,
    )

    if search:
        from app.services.search_service import tr_contains

        phone_digits = search.replace(" ", "").replace("-", "")
        query = query.where(
            db.or_(
                tr_contains(Party.name, search),
                tr_contains(Party.email, search),
                db.func.replace(db.func.replace(Party.phone, " ", ""), "-", "").ilike(
                    f"%{phone_digits}%"
                ),
            )
        )

    from app.services.search_service import tr_order

    query = query.order_by(tr_order(Party.name))
    pagination = db.paginate(query, page=max(request.args.get("page", 1, type=int), 1), per_page=25, max_per_page=100, error_out=False)
    parties = pagination.items

    # Canlı arama yalnızca sonuç tablosunu ister; sayfa yeniden yüklenmez.
    template = (
        "parties/_results.html"
        if request.args.get("partial") == "1"
        else "parties/list.html"
    )
    return render_template(
        template,
        parties=parties,
        search=search,
        pagination=pagination,
    )


@parties_bp.route("/work-orders")
@login_required
@permissions_required("clinical.view")
def list_work_orders():
    """Daily/monthly operational ledger for all doctors' work orders."""
    from app.services.validation_service import parse_date

    today = date.today()
    view = request.args.get("view", "day")
    if view == "month":
        year = request.args.get("year", today.year, type=int)
        month = request.args.get("month", today.month, type=int)
        try:
            period_start = date(year, month, 1)
        except ValueError:
            year, month = today.year, today.month
            period_start = date(year, month, 1)
        period_end = date(year + (month == 12), month % 12 + 1, 1)
        period_label = f"{MONTHS[month - 1][1]} {year}"
        selected_date = None
    else:
        view = "day"
        selected_date = parse_date(request.args.get("date", "")) or today
        period_start = selected_date
        period_end = selected_date + timedelta(days=1)
        year, month = selected_date.year, selected_date.month
        period_label = selected_date.strftime("%d.%m.%Y")

    work_orders = db.session.execute(
        db.select(WorkOrder)
        .join(Party, WorkOrder.party_id == Party.id)
        .where(
            WorkOrder.work_date >= period_start,
            WorkOrder.work_date < period_end,
            Party.is_active == True,
        )
        .order_by(WorkOrder.work_date.desc(), WorkOrder.id.desc())
    ).scalars().all()

    return render_template(
        "work_orders/list.html",
        work_orders=work_orders,
        view=view,
        selected_date=selected_date,
        year=year,
        month=month,
        months=MONTHS,
        years=range(today.year - 3, today.year + 2),
        period_label=period_label,
        doctor_count=len({wo.party_id for wo in work_orders}),
        period_total=sum((wo.total_price for wo in work_orders), Decimal("0.00")),
        today=today,
    )


@parties_bp.route("/add", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def add_party():
    if request.method == "POST":
        from app.services.validation_service import normalize_display_name

        party = Party(
            party_type=PartyType.DENTIST,
            name=normalize_display_name(request.form.get("name", "")),
            phone=request.form.get("phone", "").strip() or None,
            email=request.form.get("email", "").strip() or None,
            address=request.form.get("address", "").strip() or None,
            tax_id=request.form.get("tax_id", "").strip() or None,
            notes=request.form.get("notes", "").strip() or None,
            is_active=request.form.get("is_active") == "on",
        )
        db.session.add(party)
        db.session.commit()
        flash(f"{party.display_name} başarıyla eklendi.", "success")
        return redirect(url_for("parties.detail_party", party_id=party.id))

    return render_template("parties/form.html", party=None)


@parties_bp.route("/import", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def import_parties():
    """Diş hekimlerini Excel dosyasından toplu içe aktar (A: Ad, B: Telefon, C: E-posta, D: Adres, E: Vergi/TC No, F: Not)."""
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Dosya seçilmedi.", "danger")
            return redirect(url_for("parties.import_parties"))

        filename = (file.filename or "").lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            flash("Yalnızca .xlsx veya .xls dosyaları desteklenir.", "danger")
            return redirect(url_for("parties.import_parties"))

        try:
            import io
            from openpyxl import load_workbook
            from app.services.validation_service import normalize_display_name

            wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active
            added, updated, skipped = 0, 0, 0

            def _text(value, limit):
                return str(value).strip()[:limit] if value is not None and str(value).strip() else None

            for row in ws.iter_rows(min_row=2, values_only=True):
                name = _text(row[0] if row else None, 200)
                if not name:
                    skipped += 1
                    continue
                name = normalize_display_name(name)
                phone = _text(row[1] if len(row) > 1 else None, 20)
                email = _text(row[2] if len(row) > 2 else None, 200)
                address = _text(row[3] if len(row) > 3 else None, 2000)
                tax_id = _text(row[4] if len(row) > 4 else None, 50)
                notes = _text(row[5] if len(row) > 5 else None, 2000)

                # Yalnızca isimle eşleştir (Türkçe duyarlı): aynı telefonu paylaşan
                # farklı isimli kayıtlar (ör. iki şubeli klinik) ayrı satır kalmalı.
                from app.services.search_service import tr_equals

                existing = db.session.execute(
                    db.select(Party).where(
                        Party.party_type == PartyType.DENTIST,
                        tr_equals(Party.name, name),
                    )
                ).scalars().first()

                if existing:
                    existing.name = name
                    existing.phone = phone or existing.phone
                    existing.email = email or existing.email
                    existing.address = address or existing.address
                    existing.tax_id = tax_id or existing.tax_id
                    existing.notes = notes or existing.notes
                    existing.is_active = True
                    updated += 1
                else:
                    db.session.add(Party(
                        party_type=PartyType.DENTIST, name=name, phone=phone,
                        email=email, address=address, tax_id=tax_id, notes=notes,
                        is_active=True,
                    ))
                    added += 1

            db.session.commit()
            wb.close()
            flash(f"İçe aktarma tamamlandı: {added} yeni, {updated} güncellendi, {skipped} atlandı.", "success")
            return redirect(url_for("parties.list_parties"))
        except Exception as e:
            db.session.rollback()
            flash(f"İçe aktarma hatası: {str(e)}", "danger")
            return redirect(url_for("parties.import_parties"))

    return render_template("parties/import.html")


@parties_bp.route("/<int:party_id>")
@login_required
@permissions_required("clinical.view")
def detail_party(party_id):
    party = db.get_or_404(Party, party_id)

    work_orders = db.session.execute(
        db.select(WorkOrder)
        .where(WorkOrder.party_id == party_id)
        .order_by(WorkOrder.work_date.desc(), WorkOrder.id.desc())
    ).scalars().all()

    total_apparatus = sum(wo.apparatus_price for wo in work_orders)
    total_extra = sum(wo.extra_price for wo in work_orders)
    total_overall = sum(wo.total_price for wo in work_orders)
    total_vat = sum(db.session.execute(
        db.select(Makbuz.vat_amount).where(Makbuz.party_id == party_id)
    ).scalars().all(), Decimal("0.00"))
    total_with_vat = total_overall + total_vat

    return render_template(
        "parties/detail.html",
        party=party,
        work_orders=work_orders,
        total_apparatus=total_apparatus,
        total_extra=total_extra,
        total_overall=total_overall,
        total_vat=total_vat,
        total_with_vat=total_with_vat,
    )


@parties_bp.route("/<int:party_id>/edit", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def edit_party(party_id):
    party = db.get_or_404(Party, party_id)

    if request.method == "POST":
        from app.services.validation_service import normalize_display_name

        party.name = normalize_display_name(request.form.get("name", ""))
        party.phone = request.form.get("phone", "").strip() or None
        party.email = request.form.get("email", "").strip() or None
        party.address = request.form.get("address", "").strip() or None
        party.tax_id = request.form.get("tax_id", "").strip() or None
        party.notes = request.form.get("notes", "").strip() or None
        party.is_active = request.form.get("is_active") == "on"

        db.session.commit()
        flash(f"{party.display_name} güncellendi.", "success")
        return redirect(url_for("parties.detail_party", party_id=party.id))

    return render_template("parties/form.html", party=party)


@parties_bp.route("/<int:party_id>/delete", methods=["POST"])
@login_required
@permissions_required("clinical.delete")
def delete_party(party_id):
    party = db.get_or_404(Party, party_id)
    party.is_active = False
    db.session.commit()
    flash(f"{party.display_name} silindi.", "warning")
    return redirect(url_for("parties.list_parties"))


@parties_bp.route("/<int:party_id>/work-orders/add", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def add_work_order(party_id):
    party = db.get_or_404(Party, party_id)

    if request.method == "POST":
        from app.services.validation_service import normalize_display_name, parse_date, parse_float

        work_date = parse_date(request.form.get("work_date", ""))
        if not work_date:
            flash("Geçersiz tarih.", "danger")
            return redirect(url_for("parties.add_work_order", party_id=party.id))

        apparatus_price = parse_float(request.form.get("apparatus_price", "0")) or 0
        extra_price = parse_float(request.form.get("extra_price", "0")) or 0
        exchange_rate_applied = parse_float(request.form.get("exchange_rate_applied", "")) or None

        wo = WorkOrder(
            party_id=party_id,
            work_date=work_date,
            apparatus_type=request.form.get("apparatus_type", "").strip(),
            extra_addons=request.form.get("extra_addons", "").strip() or None,
            patient_name=normalize_display_name(request.form.get("patient_name", "")),
            apparatus_price=apparatus_price,
            extra_price=extra_price,
            total_price=apparatus_price + extra_price,
            exchange_rate_applied=exchange_rate_applied,
            notes=request.form.get("notes", "").strip() or None,
        )
        db.session.add(wo)
        db.session.commit()
        flash("İş emri başarıyla eklendi.", "success")
        return redirect(url_for("parties.detail_party", party_id=party.id))

    today = date.today().isoformat()
    from app.services.exchange_service import get_latest_rate, get_latest_usd_rate
    return render_template(
        "parties/work_order_form.html",
        party=party,
        work_order=None,
        today=today,
        ana_islemler_treatments=_get_treatments_by_category(TreatmentCategory.ANA_ISLEMLER),
        ekstra_islemler_treatments=_get_treatments_by_category(TreatmentCategory.EKSTRA_ISLEMLER),
        eur_to_try=_get_current_rate(),
        usd_to_try=get_latest_usd_rate() or 0,
    )


@parties_bp.route("/<int:party_id>/work-orders/<int:wo_id>/edit", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def edit_work_order(party_id, wo_id):
    party = db.get_or_404(Party, party_id)
    wo = db.get_or_404(WorkOrder, wo_id)

    if request.method == "POST":
        from app.services.validation_service import normalize_display_name, parse_date, parse_float

        work_date = parse_date(request.form.get("work_date", ""))
        if not work_date:
            flash("Geçersiz tarih.", "danger")
            return redirect(url_for("parties.edit_work_order", party_id=party.id, wo_id=wo.id))

        apparatus_price = parse_float(request.form.get("apparatus_price", "0")) or 0
        extra_price = parse_float(request.form.get("extra_price", "0")) or 0
        exchange_rate_applied = parse_float(request.form.get("exchange_rate_applied", "")) or None

        wo.work_date = work_date
        wo.apparatus_type = request.form.get("apparatus_type", "").strip()
        wo.extra_addons = request.form.get("extra_addons", "").strip() or None
        wo.patient_name = normalize_display_name(request.form.get("patient_name", ""))
        wo.apparatus_price = apparatus_price
        wo.extra_price = extra_price
        wo.total_price = apparatus_price + extra_price
        wo.exchange_rate_applied = exchange_rate_applied
        wo.notes = request.form.get("notes", "").strip() or None

        db.session.commit()
        flash("İş emri güncellendi.", "success")
        if request.form.get("return_to") == "work_orders":
            view = request.form.get("return_view", "day")
            if view == "month":
                return redirect(url_for(
                    "parties.list_work_orders", view="month",
                    year=request.form.get("return_year", type=int),
                    month=request.form.get("return_month", type=int),
                ))
            return redirect(url_for(
                "parties.list_work_orders", view="day",
                date=request.form.get("return_date", ""),
            ))
        return redirect(url_for("parties.detail_party", party_id=party.id))

    from app.services.exchange_service import get_latest_rate, get_latest_usd_rate
    return render_template(
        "parties/work_order_form.html",
        party=party,
        work_order=wo,
        today=wo.work_date.isoformat(),
        ana_islemler_treatments=_get_treatments_by_category(TreatmentCategory.ANA_ISLEMLER),
        ekstra_islemler_treatments=_get_treatments_by_category(TreatmentCategory.EKSTRA_ISLEMLER),
        eur_to_try=_get_current_rate(),
        usd_to_try=get_latest_usd_rate() or 0,
    )


@parties_bp.route("/<int:party_id>/work-orders/<int:wo_id>/delete", methods=["POST"])
@login_required
@permissions_required("clinical.delete")
def delete_work_order(party_id, wo_id):
    wo = db.get_or_404(WorkOrder, wo_id)
    db.session.delete(wo)
    db.session.commit()
    flash("İş emri silindi.", "warning")
    if request.form.get("return_to") == "work_orders":
        view = request.form.get("return_view", "day")
        if view == "month":
            return redirect(url_for(
                "parties.list_work_orders",
                view="month",
                year=request.form.get("return_year", type=int),
                month=request.form.get("return_month", type=int),
            ))
        return redirect(url_for(
            "parties.list_work_orders",
            view="day",
            date=request.form.get("return_date", ""),
        ))
    return redirect(url_for("parties.detail_party", party_id=party_id))
