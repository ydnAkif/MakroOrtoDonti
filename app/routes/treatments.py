from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required
import io

from app.extensions import db
from app.models.models import Treatment, TreatmentCategory
from app.authz import permissions_required
from app.services.validation_service import (
    TREATMENT_CATEGORY_ALIASES,
    normalize_treatment_fields,
)

treatments_bp = Blueprint("treatments", __name__)


def _treatment_form_values():
    """Validate and normalize the treatment fields shared by add/edit."""
    return normalize_treatment_fields(
        request.form.get("name"),
        request.form.get("description"),
        request.form.get("category"),
        request.form.get("price_eur"),
        request.form.get("currency"),
    )


@treatments_bp.route("/")
@login_required
@permissions_required("clinical.view")
def list_treatments():
    category = request.args.get("category", "")
    search = request.args.get("search", "").strip()

    query = db.select(Treatment).where(Treatment.is_active == True)

    if category:
        query = query.where(Treatment.category == category)
    if search:
        from app.services.search_service import tr_contains

        query = query.where(tr_contains(Treatment.name, search))

    query = query.order_by(Treatment.category, Treatment.name)
    pagination = db.paginate(query, page=max(request.args.get("page", 1, type=int), 1), per_page=30, max_per_page=100, error_out=False)
    treatments = pagination.items

    categories = TreatmentCategory.ALL
    category_labels = {
        "ana_islemler": "Ana İşlemler",
        "ekstra_islemler": "Ekstra İşlemler",
    }

    return render_template(
        "treatments/list.html",
        treatments=treatments,
        categories=categories,
        category_labels=category_labels,
        selected_category=category,
        search=search,
        pagination=pagination,
    )


@treatments_bp.route("/add", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def add_treatment():
    if request.method == "POST":
        try:
            name, description, category, price_eur, currency = _treatment_form_values()
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("treatments.add_treatment"))
        treatment = Treatment(
            name=name,
            description=description,
            category=category,
            price_eur=price_eur,
            currency=currency,
        )
        db.session.add(treatment)
        db.session.commit()
        flash(f"{treatment.name} eklendi.", "success")
        return redirect(url_for("treatments.list_treatments"))

    category_labels = {
        "ana_islemler": "Ana İşlemler",
        "ekstra_islemler": "Ekstra İşlemler",
    }
    return render_template("treatments/form.html", treatment=None, category_labels=category_labels)


@treatments_bp.route("/<int:treatment_id>/edit", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def edit_treatment(treatment_id):
    treatment = db.get_or_404(Treatment, treatment_id)

    if request.method == "POST":
        try:
            name, description, category, price_eur, currency = _treatment_form_values()
        except ValueError as exc:
            flash(str(exc), "danger")
            return redirect(url_for("treatments.edit_treatment", treatment_id=treatment.id))
        treatment.name = name
        treatment.description = description
        treatment.category = category
        treatment.price_eur = price_eur
        treatment.currency = currency
        db.session.commit()
        flash(f"{treatment.name} güncellendi.", "success")
        return redirect(url_for("treatments.list_treatments"))

    category_labels = {
        "ana_islemler": "Ana İşlemler",
        "ekstra_islemler": "Ekstra İşlemler",
    }
    return render_template("treatments/form.html", treatment=treatment, category_labels=category_labels)


@treatments_bp.route("/<int:treatment_id>/delete", methods=["POST"])
@login_required
@permissions_required("clinical.edit")
def delete_treatment(treatment_id):
    treatment = db.get_or_404(Treatment, treatment_id)
    treatment.is_active = False
    db.session.commit()
    flash(f"{treatment.name} silindi.", "warning")
    return redirect(url_for("treatments.list_treatments"))


CATEGORY_MAP = TREATMENT_CATEGORY_ALIASES

CATEGORY_LABELS_REV = {
    "ana_islemler": "ana_islemler", "ana işlemler": "ana_islemler",
    "ana islemler": "ana_islemler", "ana": "ana_islemler",
    "ekstra_islemler": "ekstra_islemler", "ekstra işlemler": "ekstra_islemler",
    "ekstra islemler": "ekstra_islemler", "ekstra": "ekstra_islemler",
}


@treatments_bp.route("/import", methods=["GET", "POST"])
@login_required
@permissions_required("clinical.edit")
def import_treatments():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Dosya seçilmedi.", "danger")
            return redirect(url_for("treatments.import_treatments"))

        filename = file.filename.lower()
        if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
            flash("Yalnızca .xlsx veya .xls dosyaları desteklenir.", "danger")
            return redirect(url_for("treatments.import_treatments"))

        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(min_row=2, values_only=True))
            added = 0
            updated = 0
            skipped = 0

            for row in rows:
                if not row or not row[0]:
                    skipped += 1
                    continue

                try:
                    name, description, category, price, currency = normalize_treatment_fields(
                        row[0],
                        row[3] if len(row) > 3 else None,
                        row[1] if len(row) > 1 else "other",
                        row[2] if len(row) > 2 else None,
                        row[4] if len(row) > 4 else None,
                    )
                except ValueError:
                    skipped += 1
                    continue

                existing = db.session.execute(
                    db.select(Treatment).where(Treatment.name == name)
                ).scalar_one_or_none()

                if existing:
                    existing.price_eur = price
                    existing.category = category
                    existing.currency = currency
                    if description:
                        existing.description = description
                    updated += 1
                else:
                    db.session.add(Treatment(
                        name=name,
                        category=category,
                        price_eur=price,
                        currency=currency,
                        description=description,
                        is_active=True,
                    ))
                    added += 1

            db.session.commit()
            wb.close()

            parts = []
            if added:
                parts.append(f"{added} yeni")
            if updated:
                parts.append(f"{updated} güncellendi")
            if skipped:
                parts.append(f"{skipped} atlandı")

            flash(f"İçe aktarma tamamlandı: {', '.join(parts) if parts else 'Değişiklik yok'}.", "success")
        except Exception as e:
            flash(f"İçe aktarma hatası: {str(e)}", "danger")

        return redirect(url_for("treatments.list_treatments"))

    return render_template("treatments/import.html")


@treatments_bp.route("/api/update", methods=["POST"])
@login_required
@permissions_required("clinical.edit")
def api_update_treatment():
    data = request.get_json()
    if not data or "id" not in data:
        return jsonify({"error": "Geçersiz istek"}), 400

    treatment = db.session.get(Treatment, data["id"])
    if not treatment:
        return jsonify({"error": "İşlem bulunamadı"}), 404

    try:
        name, description, category, price, currency = normalize_treatment_fields(
            data.get("name", treatment.name),
            data.get("description", treatment.description),
            data.get("category", treatment.category),
            data.get("price_eur", treatment.price_eur),
            data.get("currency", treatment.currency),
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    treatment.name = name
    treatment.description = description
    treatment.category = category
    treatment.price_eur = price
    treatment.currency = currency

    db.session.commit()

    return jsonify({
        "ok": True,
        "treatment": {
            "id": treatment.id,
            "name": treatment.name,
            "category": treatment.category,
            "price_eur": float(treatment.price_eur),
            "currency": treatment.currency,
            "description": treatment.description or "",
        }
    })
